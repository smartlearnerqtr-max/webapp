from __future__ import annotations

import csv
import io
import secrets
import string
import unicodedata

from flask import request, send_file
from flask_jwt_extended import get_jwt, get_jwt_identity, jwt_required

from ....extensions import db
from ....models import (
    Classroom,
    ClassStudent,
    LessonAssignmentStudent,
    ParentDailyReport,
    ParentStudentLink,
    ParentTeacherMessage,
    RealtimeEvent,
    ServerLog,
    StudentAccountBatch,
    StudentAccountBatchMember,
    StudentLessonProgress,
    StudentProfile,
    TeacherParentStudentLink,
    TeacherStudentLink,
    User,
    UserAISetting,
)
from ....services.auth_service import create_teacher_user
from ....services.github_snapshot_service import JsonSnapshotPersistError, persist_json_snapshot
from ....services.logger import log_server_event
from ....utils.security import hash_password
from ....utils.responses import error_response, success_response
from .. import api_v1


def _require_admin_user():
    if get_jwt().get('role') != 'admin':
        return None, error_response('Không có quyền truy cập', 'AUTH_FORBIDDEN', 403)
    user = User.query.get(get_jwt_identity())
    if not user or user.role != 'admin':
        return None, error_response('Không tìm thấy admin', 'ADMIN_NOT_FOUND', 404)
    return user, None


def _persist_json_snapshot_or_error(action_name: str):
    try:
        persist_json_snapshot(action_name)
    except JsonSnapshotPersistError as error:
        db.session.rollback()
        return error_response(str(error), 'JSON_SNAPSHOT_PERSIST_FAILED', 502, error.details)
    return None


def _build_teacher_payload(teacher_user: User) -> dict[str, object]:
    return {
        'user': teacher_user.to_dict(),
        'profile': teacher_user.teacher_profile.to_dict() if teacher_user.teacher_profile else None,
    }


def _account_username(user: User) -> str | None:
    return user.email or user.phone


def _build_recovery_account_payload(user: User) -> dict[str, object]:
    profile = user.teacher_profile if user.role == 'teacher' else user.student_profile
    profile_payload = profile.to_dict() if profile else None
    login_id = user.phone or str(user.id)
    return {
        'user': user.to_dict(),
        'profile': profile_payload,
        'full_name': profile_payload.get('full_name') if profile_payload else None,
        'login_id': login_id,
        'username': _account_username(user),
        'can_login': bool(_account_username(user)),
    }


def _normalize_column_name(value: object) -> str:
    raw = str(value or '').strip().lower().replace('đ', 'd').replace('Đ', 'D')
    without_accents = ''.join(
        char for char in unicodedata.normalize('NFKD', raw)
        if not unicodedata.combining(char)
    )
    return ''.join(char for char in without_accents if char.isalnum())


def _pick(row: dict[str, object], aliases: set[str]) -> str:
    for key, value in row.items():
        if _normalize_column_name(key) in aliases:
            return str(value or '').strip()
    return ''


def _pick_by_index(row: dict[str, object], index: int) -> str:
    values = list(row.values())
    if index >= len(values):
        return ''
    return str(values[index] or '').strip()


def _generate_batch_code() -> str:
    alphabet = string.ascii_uppercase + string.digits
    while True:
        code = ''.join(secrets.choice(alphabet) for _ in range(9))
        if not StudentAccountBatch.query.filter_by(code=code).first():
            return code


def _is_valid_student_code(value: str) -> bool:
    return value.isdigit() and len(value) == 7


def _parse_student_account_file(uploaded_file) -> list[dict[str, object]]:
    filename = (uploaded_file.filename or '').lower()
    raw_bytes = uploaded_file.read()
    if filename.endswith('.csv'):
        text = raw_bytes.decode('utf-8-sig')
        return [dict(row) for row in csv.DictReader(io.StringIO(text))]

    if filename.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise RuntimeError('Server chưa cài openpyxl để đọc file Excel .xlsx') from error

        workbook = load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(header or '').strip() for header in rows[0]]
        records: list[dict[str, object]] = []
        for values in rows[1:]:
            records.append({headers[index]: values[index] if index < len(values) else None for index in range(len(headers))})
        return records

    raise ValueError('Chỉ hỗ trợ file .xlsx hoặc .csv')


def _build_batch_payload(batch: StudentAccountBatch, include_members: bool = False) -> dict[str, object]:
    payload = batch.to_dict()
    if include_members:
        payload['members'] = [member.to_dict() for member in batch.members if member.status == 'active']
    return payload


def _level_label(level: str | None) -> str:
    if level == 'nhe':
        return 'Nhẹ'
    if level == 'trung_binh':
        return 'Trung bình'
    if level == 'nang':
        return 'Nặng'
    return level or 'Chưa rõ'


def _build_admin_class_student_payload(class_link: ClassStudent) -> dict[str, object]:
    student = class_link.student
    user = student.user if student and student.user else None
    return {
        'link_id': class_link.id,
        'student_profile_id': student.id if student else None,
        'user_id': user.id if user else None,
        'login_id': user.phone if user and user.phone else str(user.id) if user else None,
        'full_name': student.full_name if student else None,
        'disability_level': student.disability_level if student else None,
        'disability_level_label': _level_label(student.disability_level if student else None),
        'username': user.email or user.phone if user else None,
        'email': user.email if user else None,
        'phone': user.phone if user else None,
        'account_status': user.status if user else 'missing',
        'class_link_status': class_link.status,
        'support_note': student.support_note if student else None,
        'preferred_input': student.preferred_input if student else None,
    }


def _build_admin_class_payload(classroom: Classroom) -> dict[str, object]:
    active_student_links = [link for link in classroom.students if link.status == 'active']
    level_counts: dict[str, int] = {}
    for link in active_student_links:
        level = link.student.disability_level if link.student else None
        level_counts[_level_label(level)] = level_counts.get(_level_label(level), 0) + 1

    teacher = classroom.teacher
    teacher_user = teacher.user if teacher and teacher.user else None
    payload = classroom.to_dict(include_join_credentials=True)
    payload.update({
        'teacher': {
            'id': teacher.id if teacher else None,
            'user_id': teacher_user.id if teacher_user else None,
            'full_name': teacher.full_name if teacher else None,
            'school_name': teacher.school_name if teacher else None,
            'email': teacher_user.email if teacher_user else None,
            'phone': teacher_user.phone if teacher_user else None,
        } if teacher else None,
        'student_count': len(active_student_links),
        'level_counts': level_counts,
        'students': [_build_admin_class_student_payload(link) for link in active_student_links],
    })
    return payload


def _excel_filename(value: str) -> str:
    normalized = unicodedata.normalize('NFKD', value or 'lop-hoc')
    ascii_text = ''.join(char for char in normalized if not unicodedata.combining(char)).lower()
    cleaned = ''.join(char if char.isalnum() else '-' for char in ascii_text)
    cleaned = '-'.join(part for part in cleaned.split('-') if part)
    return cleaned or 'lop-hoc'


def _build_classes_excel(classrooms: list[Classroom]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Danh sách lớp'

    title_fill = PatternFill('solid', fgColor='1F4E79')
    title_font = Font(color='FFFFFF', bold=True, size=15)
    section_fill = PatternFill('solid', fgColor='D9EAF7')
    header_fill = PatternFill('solid', fgColor='EAF2F8')
    header_font = Font(bold=True, color='1F2937')
    thin_border = Border(
        left=Side(style='thin', color='D9E2EC'),
        right=Side(style='thin', color='D9E2EC'),
        top=Side(style='thin', color='D9E2EC'),
        bottom=Side(style='thin', color='D9E2EC'),
    )

    columns = [
        ('STT', 7),
        ('ID đăng nhập', 16),
        ('Họ tên học sinh', 28),
        ('Tên tài khoản', 30),
        ('Email', 32),
        ('Mức độ', 16),
        ('Trạng thái', 18),
    ]
    for index, (_header, width) in enumerate(columns, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    row_index = 1
    for class_index, classroom in enumerate(classrooms, start=1):
        active_links = [link for link in classroom.students if link.status == 'active']
        teacher = classroom.teacher
        teacher_user = teacher.user if teacher and teacher.user else None

        worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(columns))
        title_cell = worksheet.cell(row=row_index, column=1, value=f'{class_index}. {classroom.name}')
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(vertical='center')
        worksheet.row_dimensions[row_index].height = 26
        row_index += 1

        meta_rows = [
            ('Khối/Lớp', classroom.grade_label or 'Chưa có'),
            ('Giáo viên', teacher.full_name if teacher else 'Chưa có giáo viên'),
            ('Email giáo viên', teacher_user.email if teacher_user and teacher_user.email else ''),
            ('SĐT giáo viên', teacher_user.phone if teacher_user and teacher_user.phone else ''),
            ('Mã lớp', classroom.join_credential.join_password if classroom.join_credential else ''),
            ('Mức độ lớp', _level_label(classroom.default_disability_level)),
            ('Số học sinh', len(active_links)),
        ]
        for label, value in meta_rows:
            worksheet.cell(row=row_index, column=1, value=label)
            worksheet.cell(row=row_index, column=2, value=value)
            worksheet.cell(row=row_index, column=1).font = Font(bold=True)
            worksheet.cell(row=row_index, column=1).fill = section_fill
            worksheet.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True)
            row_index += 1

        row_index += 1
        for column_index, (header, _width) in enumerate(columns, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        row_index += 1

        if active_links:
            for student_index, class_link in enumerate(active_links, start=1):
                student = class_link.student
                user = student.user if student and student.user else None
                values = [
                    student_index,
                    user.phone if user and user.phone else user.id if user else '',
                    student.full_name if student else '',
                    user.email or user.phone if user else '',
                    user.email if user and user.email else '',
                    _level_label(student.disability_level if student else None),
                    readable_status_for_excel(user.status if user else 'missing'),
                ]
                for column_index, value in enumerate(values, start=1):
                    cell = worksheet.cell(row=row_index, column=column_index, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                row_index += 1
        else:
            worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=len(columns))
            cell = worksheet.cell(row=row_index, column=1, value='Lớp này chưa có học sinh.')
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            row_index += 1

        row_index += 2

    worksheet.freeze_panes = 'A10'
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def readable_status_for_excel(status: str | None) -> str:
    if status == 'active':
        return 'Đang hoạt động'
    if status == 'inactive':
        return 'Tạm khóa'
    if status == 'archived':
        return 'Lưu trữ'
    if status == 'missing':
        return 'Thiếu tài khoản'
    return status or 'Chưa rõ'


@api_v1.get('/admin/teachers')
@jwt_required()
def list_teachers():
    user, error = _require_admin_user()
    if error:
        return error
    query = User.query.filter_by(role='teacher')
    if request.args.get('status'):
        query = query.filter_by(status=request.args['status'])
    teachers = query.order_by(User.created_at.desc()).all()
    payload = [_build_teacher_payload(teacher) for teacher in teachers]
    return success_response(payload)


@api_v1.get('/admin/classes/overview')
@jwt_required()
def list_admin_classes_overview():
    _user, error = _require_admin_user()
    if error:
        return error

    classrooms = Classroom.query.order_by(Classroom.created_at.desc()).all()
    payload = [_build_admin_class_payload(classroom) for classroom in classrooms]
    for classroom in payload:
        classroom['students'] = sorted(
            classroom['students'],
            key=lambda student: str(student.get('full_name') or '').lower(),
        )
    return success_response({
        'summary': {
            'class_count': len(payload),
            'student_count': sum(int(classroom['student_count']) for classroom in payload),
        },
        'classes': payload,
    })


@api_v1.get('/admin/classes/overview/export')
@jwt_required()
def export_admin_classes_overview():
    _user, error = _require_admin_user()
    if error:
        return error

    class_id = request.args.get('class_id')
    query = Classroom.query.order_by(Classroom.created_at.desc())
    filename = 'danh-sach-lop-hoc'
    if class_id:
        try:
            class_id_value = int(class_id)
        except (TypeError, ValueError):
            return error_response('class_id không hợp lệ', 'VALIDATION_ERROR', 422)
        query = query.filter_by(id=class_id_value)

    classrooms = query.all()
    if class_id and not classrooms:
        return error_response('Không tìm thấy lớp', 'CLASS_NOT_FOUND', 404)
    if len(classrooms) == 1:
        filename = f"danh-sach-{_excel_filename(classrooms[0].name)}"

    output = _build_classes_excel(classrooms)
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'{filename}.xlsx',
        max_age=0,
    )


@api_v1.get('/admin/account-recovery')
@jwt_required()
def list_recoverable_accounts():
    _user, error = _require_admin_user()
    if error:
        return error

    query = User.query.filter(User.role.in_(('teacher', 'student')))
    role = (request.args.get('role') or '').strip()
    status = (request.args.get('status') or '').strip()
    if role in {'teacher', 'student'}:
        query = query.filter_by(role=role)
    if status:
        query = query.filter_by(status=status)

    accounts = query.order_by(User.role.asc(), User.created_at.desc()).all()
    return success_response([_build_recovery_account_payload(account) for account in accounts])


@api_v1.get('/admin/student-account-batches')
@jwt_required()
def list_student_account_batches():
    _user, error = _require_admin_user()
    if error:
        return error

    batches = StudentAccountBatch.query.order_by(StudentAccountBatch.created_at.desc()).all()
    return success_response([_build_batch_payload(batch) for batch in batches])


@api_v1.post('/admin/student-account-batches/import')
@jwt_required()
def import_student_account_batch():
    admin_user, error = _require_admin_user()
    if error:
        return error

    uploaded_file = request.files.get('file')
    if not uploaded_file:
        return error_response('Vui lòng chọn file Excel hoặc CSV', 'VALIDATION_ERROR', 422)

    batch_title = (request.form.get('title') or uploaded_file.filename or 'Danh sách học sinh').strip()
    try:
        records = _parse_student_account_file(uploaded_file)
    except (ValueError, RuntimeError) as parse_error:
        return error_response(str(parse_error), 'VALIDATION_ERROR', 422)

    student_code_aliases = {'id', 'ma', 'mahocsinh', 'studentid', 'studentcode', 'code', 'madangnhap', 'mshs'}
    full_name_aliases = {'ten', 'hoten', 'hovaten', 'tenhocsinh', 'hotenhocsinh', 'fullname', 'name', 'studentname'}
    username_aliases = {'tentaikhoan', 'taikhoan', 'username', 'account', 'email'}
    password_aliases = {'matkhau', 'password', 'pass'}
    level_aliases = {'mucdo', 'mucdohotro', 'disabilitylevel', 'level'}

    created_count = 0
    skipped_rows: list[dict[str, object]] = []
    imported_levels: set[str] = set()
    parsed_rows: list[dict[str, str]] = []
    seen_student_codes: set[str] = set()
    duplicate_file_ids: set[str] = set()

    for row_index, row in enumerate(records, start=2):
        student_code = (_pick(row, student_code_aliases) or _pick_by_index(row, 0)).upper()
        full_name = _pick(row, full_name_aliases) or _pick_by_index(row, 1)
        username = _pick(row, username_aliases) or _pick_by_index(row, 2)
        password = _pick(row, password_aliases) or _pick_by_index(row, 3) or 'Student123!'
        level = _pick(row, level_aliases) or _pick_by_index(row, 4) or 'trung_binh'
        if level not in {'nhe', 'trung_binh', 'nang'}:
            level = 'trung_binh'

        if not student_code or not full_name:
            skipped_rows.append({'row': row_index, 'reason': 'Thiếu mã học sinh hoặc họ tên'})
            continue
        if not _is_valid_student_code(student_code):
            skipped_rows.append({'row': row_index, 'reason': 'Mã học sinh phải gồm đúng 7 chữ số, ví dụ 2026001'})
            continue
        if student_code in seen_student_codes:
            duplicate_file_ids.add(student_code)
        seen_student_codes.add(student_code)
        imported_levels.add(level)
        parsed_rows.append({
            'student_code': student_code,
            'full_name': full_name,
            'username': username,
            'password': password,
            'level': level,
        })

    if duplicate_file_ids:
        return error_response(
            'File có ID bị trùng. Vui lòng cấp ID khác cho các học sinh này.',
            'DUPLICATE_STUDENT_IDS_IN_FILE',
            422,
            {'conflicting_ids': sorted(duplicate_file_ids)},
        )

    existing_student_codes = {
        code
        for (code,) in db.session.query(User.phone)
        .filter(User.phone.in_([row['student_code'] for row in parsed_rows]))
        .all()
        if code
    }
    if existing_student_codes:
        return error_response(
            'ID đã tồn tại. Vui lòng cấp ID khác.',
            'STUDENT_ID_ALREADY_EXISTS',
            409,
            {'conflicting_ids': sorted(existing_student_codes)},
        )

    if len(imported_levels) > 1:
        return error_response('Một danh sách học sinh chỉ được có một mức độ: nhẹ, trung bình hoặc nặng. Vui lòng tách file theo từng mức độ.', 'MIXED_LEVEL_BATCH', 422)

    batch = StudentAccountBatch(
        code=_generate_batch_code(),
        title=batch_title,
        created_by_admin_id=admin_user.id,
        status='active',
    )
    db.session.add(batch)
    db.session.flush()

    for row in parsed_rows:
        student_code = row['student_code']
        username = row['username']
        password = row['password']
        email = username.lower() if '@' in username else None
        if email and User.query.filter_by(email=email).first():
            email = None
        user = User(
            email=email,
            phone=student_code,
            password_hash=hash_password(password),
            role='student',
            status='active',
        )
        db.session.add(user)
        db.session.flush()
        created_count += 1

        student_profile = StudentProfile(
            user_id=user.id,
            full_name=row['full_name'],
            disability_level=row['level'],
            support_note=f'Tài khoản import từ lô {batch.code}',
            preferred_input='touch',
        )
        db.session.add(student_profile)
        db.session.flush()

        db.session.add(StudentAccountBatchMember(
            batch_id=batch.id,
            student_id=student_profile.id,
            user_id=user.id,
            student_code=student_code,
            username=username or student_code,
            temporary_password=password,
            status='active',
        ))

    db.session.flush()
    persist_error = _persist_json_snapshot_or_error('admin_import_student_batch')
    if persist_error:
        return persist_error
    db.session.commit()

    log_server_event(
        level='info',
        module='admin',
        message='Admin import danh sách tài khoản học sinh',
        action_name='admin_import_student_batch',
        user_id=admin_user.id,
        metadata={'batch_id': batch.id, 'batch_code': batch.code, 'created_count': created_count, 'updated_count': 0},
    )

    refreshed_batch = StudentAccountBatch.query.get(batch.id)

    return success_response({
        'batch': _build_batch_payload(refreshed_batch or batch, include_members=True),
        'created_count': created_count,
        'updated_count': 0,
        'skipped_rows': skipped_rows,
    }, 'Import danh sách học sinh thành công', 201)


@api_v1.get('/admin/relationships/overview')
@jwt_required()
def get_relationship_overview():
    user, error = _require_admin_user()
    if error:
        return error

    teachers = User.query.filter_by(role='teacher').order_by(User.created_at.desc()).all()
    active_teacher_student_links = TeacherStudentLink.query.filter_by(status='active').all()
    active_parent_group_links = TeacherParentStudentLink.query.filter_by(status='active').all()

    teacher_payload = []
    for teacher in teachers:
        teacher_profile = teacher.teacher_profile
        if not teacher_profile:
            continue
        student_links = [link for link in active_teacher_student_links if link.teacher_id == teacher_profile.id]
        parent_links = [link for link in active_parent_group_links if link.teacher_id == teacher_profile.id]
        shared_student_count = len([
            link for link in student_links
            if len([peer for peer in active_teacher_student_links if peer.student_id == link.student_id]) > 1
        ])
        teacher_payload.append({
            'teacher': _build_teacher_payload(teacher),
            'student_count': len(student_links),
            'parent_group_count': len(parent_links),
            'shared_student_count': shared_student_count,
        })

    shared_students_map: dict[int, dict[str, object]] = {}
    for link in active_teacher_student_links:
        if not link.student or not link.teacher:
            continue
        entry = shared_students_map.setdefault(link.student_id, {
            'student': link.student.to_dict(),
            'teachers': [],
        })
        entry['teachers'].append({
            'id': link.teacher.id,
            'full_name': link.teacher.full_name,
            'school_name': link.teacher.school_name,
        })

    shared_students = [item for item in shared_students_map.values() if len(item['teachers']) > 1]
    shared_students.sort(key=lambda item: str(item['student']['full_name']).lower())

    return success_response({
        'summary': {
            'teacher_count': len(teacher_payload),
            'teacher_student_link_count': len(active_teacher_student_links),
            'teacher_parent_group_count': len(active_parent_group_links),
            'shared_student_count': len(shared_students),
        },
        'teachers': teacher_payload,
        'shared_students': shared_students,
    })


@api_v1.post('/admin/teachers')
@jwt_required()
def create_teacher():
    user, error = _require_admin_user()
    if error:
        return error

    payload = request.get_json(silent=True) or {}
    created_payload, error_message, error_code = create_teacher_user(payload)
    if error_message or not created_payload:
        status_code = 422 if error_code == 'VALIDATION_ERROR' else 409
        if error_code == 'JSON_SNAPSHOT_PERSIST_FAILED':
            status_code = 502
        return error_response(error_message or 'Không tạo được tài khoản giáo viên', error_code or 'VALIDATION_ERROR', status_code)

    log_server_event(level='info', module='admin', message='Admin tạo tài khoản giáo viên', action_name='admin_create_teacher', user_id=user.id, metadata={'teacher_user_id': created_payload['user']['id']})
    return success_response(created_payload, 'Tạo tài khoản giáo viên thành công', 201)


@api_v1.post('/admin/users/<int:user_id>/recover')
@jwt_required()
def recover_account(user_id: int):
    admin_user, error = _require_admin_user()
    if error:
        return error

    target_user = User.query.get(user_id)
    if not target_user or target_user.role not in {'teacher', 'student'}:
        return error_response('Không tìm thấy tài khoản giáo viên hoặc học sinh', 'USER_NOT_FOUND', 404)

    username = _account_username(target_user)
    if not username:
        return error_response('Tài khoản này chưa có email hoặc số điện thoại để đăng nhập', 'USERNAME_MISSING', 422)

    payload = request.get_json(silent=True) or {}
    temporary_password = (payload.get('temporary_password') or 'Demo123456').strip()
    if len(temporary_password) < 6:
        return error_response('Mật khẩu tạm thời cần tối thiểu 6 ký tự', 'VALIDATION_ERROR', 422)

    target_user.password_hash = hash_password(temporary_password)
    target_user.status = 'active'
    db.session.flush()
    persist_error = _persist_json_snapshot_or_error('admin_recover_account')
    if persist_error:
        return persist_error
    db.session.commit()

    log_server_event(
        level='warning',
        module='admin',
        message='Admin khôi phục mật khẩu tài khoản',
        action_name='admin_recover_account',
        user_id=admin_user.id,
        metadata={'target_user_id': target_user.id, 'target_role': target_user.role},
    )
    return success_response({
        'account': _build_recovery_account_payload(target_user),
        'username': username,
        'temporary_password': temporary_password,
    }, 'Khôi phục tài khoản thành công')


@api_v1.delete('/admin/users/<int:user_id>')
@jwt_required()
def delete_student_account(user_id: int):
    admin_user, error = _require_admin_user()
    if error:
        return error

    target_user = User.query.get(user_id)
    if not target_user or target_user.role != 'student':
        return error_response('Chỉ có thể xóa tài khoản học sinh ở màn hình này', 'STUDENT_USER_NOT_FOUND', 404)

    deleted_account = _build_recovery_account_payload(target_user)
    student_profile = target_user.student_profile
    deleted_counts: dict[str, int] = {}

    deleted_counts['realtime_events'] = RealtimeEvent.query.filter_by(recipient_user_id=target_user.id).delete(synchronize_session=False)
    deleted_counts['server_logs'] = ServerLog.query.filter_by(user_id=target_user.id).delete(synchronize_session=False)
    deleted_counts['ai_settings'] = UserAISetting.query.filter_by(user_id=target_user.id).delete(synchronize_session=False)

    if student_profile:
        student_id = student_profile.id
        deleted_counts['parent_teacher_messages'] = ParentTeacherMessage.query.filter(
            (ParentTeacherMessage.student_id == student_id) | (ParentTeacherMessage.sender_user_id == target_user.id)
        ).delete(synchronize_session=False)
        deleted_counts['lesson_progress'] = StudentLessonProgress.query.filter_by(student_id=student_id).delete(synchronize_session=False)
        deleted_counts['lesson_assignment_students'] = LessonAssignmentStudent.query.filter_by(student_id=student_id).delete(synchronize_session=False)
        deleted_counts['student_batch_members'] = StudentAccountBatchMember.query.filter(
            (StudentAccountBatchMember.student_id == student_id) | (StudentAccountBatchMember.user_id == target_user.id)
        ).delete(synchronize_session=False)
        deleted_counts['class_students'] = ClassStudent.query.filter_by(student_id=student_id).delete(synchronize_session=False)
        deleted_counts['teacher_student_links'] = TeacherStudentLink.query.filter_by(student_id=student_id).delete(synchronize_session=False)
        deleted_counts['parent_student_links'] = ParentStudentLink.query.filter_by(student_id=student_id).delete(synchronize_session=False)
        deleted_counts['teacher_parent_student_links'] = TeacherParentStudentLink.query.filter_by(student_id=student_id).delete(synchronize_session=False)
        deleted_counts['parent_daily_reports'] = ParentDailyReport.query.filter_by(student_id=student_id).delete(synchronize_session=False)
        db.session.delete(student_profile)
    else:
        deleted_counts['student_batch_members'] = StudentAccountBatchMember.query.filter_by(user_id=target_user.id).delete(synchronize_session=False)

    db.session.delete(target_user)
    db.session.flush()

    log_server_event(
        level='warning',
        module='admin',
        message='Admin xóa vĩnh viễn tài khoản học sinh',
        action_name='admin_delete_student_account',
        user_id=admin_user.id,
        metadata={'target_user_id': user_id, 'deleted_counts': deleted_counts},
    )
    db.session.flush()
    persist_error = _persist_json_snapshot_or_error('admin_delete_student_account')
    if persist_error:
        return persist_error
    db.session.commit()

    return success_response({
        'account': deleted_account,
        'deleted_counts': deleted_counts,
    }, 'Đã xóa tài khoản học sinh và toàn bộ dữ liệu liên quan')
